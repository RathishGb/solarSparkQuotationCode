"""Internal Cost Calculator & Profit Tracker for Solar Installations.

Separate from client-facing quotation PDF generator.
Calculates actual expenses, selling price, and profit margins.
Exports detailed breakdown to Excel.
"""

import math
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


# ─── Cost Constants ───
PANEL_WATT = 610  # Wp per panel

# Panel price per watt (before GST)
PANEL_RATE = {
    "domestic": 26.75,     # Rs/Wp for domestic (DCR panels)
    "commercial": 15.50,   # Rs/Wp for commercial
}
# GST rates for panel: 70% at 5%, 30% at 12%
PANEL_GST_70 = 0.05
PANEL_GST_30 = 0.18

# Inverter cost (base price before GST)
INVERTER_COST = {
    "1Phase": {
        3: 15500,
        4: 19300,
        5: 25500,
        "default": 18000,
    },
    "3Phase": 55000,
}

# Other component costs (approximate per kW)
OTHER_COSTS_PER_KW = {
    "mounting_structure": 6000,
    "dc_distribution_box": 800,
    "ac_distribution_board": 1200,
    "earthing_system": 1500,
    "cables_ac_dc": 2500,
    "monitoring_system": 500,
    # The following are now part of fixed non-GST cost:
    # "installation_labour": 3000,
    # "civil_works": 2000,
    # "transport": 1000,
    # Miscellaneous GST-applicable items can be added here
    "net_metering_liaison": 1500,
}

# Fixed non-GST cost for labor, delivery, installation, civil work
FIXED_NON_GST_COST = 30000


def calculate_costs(kw, customer_type, phase, panel_price_per_watt=None,
                    inverter_price=None, selling_price=0):
    """Calculate detailed cost breakdown for a solar installation.

    Args:
        kw: System size in kilowatts
        customer_type: 'domestic' or 'commercial'
        phase: '1Phase' or '3Phase'
        panel_price_per_watt: Override default panel rate (before GST)
        inverter_price: Override default inverter cost
        selling_price: What we quote the customer (incl GST)

    Returns:
        dict with full cost breakdown
    """
    # Panel calculations
    num_panels = math.ceil((kw * 1000) / PANEL_WATT)
    total_wp = num_panels * PANEL_WATT

    if panel_price_per_watt is None:
        panel_price_per_watt = PANEL_RATE.get(customer_type, PANEL_RATE["domestic"])


    panel_cost_before_gst = total_wp * panel_price_per_watt
    # GST: 70% at 5%, 30% at 12%
    panel_cost_70 = panel_cost_before_gst * 0.7
    panel_cost_30 = panel_cost_before_gst * 0.3
    panel_gst_70 = panel_cost_70 * PANEL_GST_70
    panel_gst_30 = panel_cost_30 * PANEL_GST_30
    panel_gst = panel_gst_70 + panel_gst_30
    panel_cost_total = panel_cost_before_gst + panel_gst


    # Inverter
    if inverter_price is None:
        if phase == "1Phase":
            # Use special rates for 3, 4, 5 kW
            kw_rounded = round(kw)
            if kw_rounded in INVERTER_COST["1Phase"]:
                inverter_price = INVERTER_COST["1Phase"][kw_rounded]
            else:
                inverter_price = INVERTER_COST["1Phase"]["default"]
        else:
            inverter_price = INVERTER_COST.get(phase, INVERTER_COST["3Phase"])
    inverter_gst = inverter_price * 0.18
    inverter_total = inverter_price + inverter_gst


    # Other costs (GST-applicable only)
    other_items = {}
    other_total = 0
    for item, rate in OTHER_COSTS_PER_KW.items():
        cost = round(rate * kw)
        other_items[item] = cost
        other_total += cost

    other_gst = round(other_total * 0.18)  # 18% GST on services
    other_total_with_gst = other_total + other_gst

    # Fixed non-GST cost
    non_gst_items = {
        "labour_delivery_installation_civil": FIXED_NON_GST_COST
    }
    non_gst_total = FIXED_NON_GST_COST


    # Totals
    total_cost_before_gst = panel_cost_before_gst + inverter_price + other_total + non_gst_total
    total_gst = panel_gst + inverter_gst + other_gst
    total_cost = panel_cost_total + inverter_total + other_total_with_gst + non_gst_total

    # Profit
    profit = selling_price - total_cost if selling_price > 0 else 0
    margin_pct = (profit / selling_price * 100) if selling_price > 0 else 0

    return {
        "kw": kw,
        "customer_type": customer_type,
        "phase": phase,
        "num_panels": num_panels,
        "total_wp": total_wp,
        "panel_price_per_watt": panel_price_per_watt,
        "panel_cost_before_gst": round(panel_cost_before_gst),
        "panel_gst_70": round(panel_gst_70),
        "panel_gst_30": round(panel_gst_30),
        "panel_gst": round(panel_gst),
        "panel_cost_total": round(panel_cost_total),
        "inverter_price": inverter_price,
        "inverter_gst": round(inverter_gst),
        "inverter_total": round(inverter_total),
        "other_items": other_items,
        "other_total_before_gst": other_total,
        "other_gst": other_gst,
        "other_total": other_total_with_gst,
        "non_gst_items": non_gst_items,
        "non_gst_total": non_gst_total,
        "total_cost_before_gst": round(total_cost_before_gst),
        "total_gst": round(total_gst),
        "total_cost": round(total_cost),
        "selling_price": selling_price,
        "profit": round(profit),
        "margin_pct": round(margin_pct, 1),
        "cost_breakup": {
            "gst_component": {
                "panel": {
                    "base": round(panel_cost_before_gst),
                    "gst_70": round(panel_gst_70),
                    "gst_30": round(panel_gst_30),
                    "total": round(panel_cost_total),
                },
                "inverter": {
                    "base": round(inverter_price),
                    "gst": round(inverter_gst),
                    "total": round(inverter_total),
                },
                "other": {
                    "base": other_total,
                    "gst": other_gst,
                    "total": other_total_with_gst,
                },
                "total_gst": round(total_gst),
            },
            "non_gst_component": {
                "items": non_gst_items,
                "total": non_gst_total,
            }
        }
    }


def generate_cost_excel(entries):
    """Generate Excel workbook with cost breakdown and profit analysis.

    Args:
        entries: list of dicts, each from calculate_costs()

    Returns:
        bytes of the Excel file
    """
    wb = Workbook()

    # ─── Styles ───
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    section_font = Font(name="Calibri", bold=True, size=11, color="1A3A5C")
    section_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
    profit_fill_green = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    profit_fill_red = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    currency_fmt = '#,##0'
    pct_fmt = '0.0"%"'

    # ────────────────────────────────────────
    # Sheet 1: Detailed Cost Breakdown
    # ────────────────────────────────────────
    ws = wb.active
    ws.title = "Cost Breakdown"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    for idx, entry in enumerate(entries):
        start_row = idx * 40 + 1  # space between entries

        # Title
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        title_cell = ws.cell(row=start_row, column=1,
                             value=f"Cost Analysis - {entry['kw']}kW {entry['customer_type'].title()} ({entry['phase']})")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1A3A5C")
        title_cell.alignment = Alignment(horizontal="center")

        # Headers
        row = start_row + 2
        for col, header in enumerate(["Component", "Amount (Rs.)", "GST (Rs.)", "Total (Rs.)"], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Panel Section
        row += 1
        ws.cell(row=row, column=1, value="SOLAR PANELS").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = border

        row += 1
        items = [
            (f"Panels ({entry['num_panels']} x {PANEL_WATT}Wp = {entry['total_wp']}Wp)",
             None, None, None),
            (f"Panel Rate: Rs.{entry['panel_price_per_watt']}/Wp + 12% GST",
             entry['panel_cost_before_gst'], entry['panel_gst'], entry['panel_cost_total']),
        ]
        for label, amt, gst, total in items:
            ws.cell(row=row, column=1, value=label).border = border
            if amt is not None:
                ws.cell(row=row, column=2, value=amt).number_format = currency_fmt
                ws.cell(row=row, column=3, value=gst).number_format = currency_fmt
                ws.cell(row=row, column=4, value=total).number_format = currency_fmt
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = border
            row += 1

        # Inverter Section
        ws.cell(row=row, column=1, value="INVERTER").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = border

        row += 1
        ws.cell(row=row, column=1, value=f"Grid Tied Inverter ({entry['phase']})").border = border
        ws.cell(row=row, column=2, value=entry['inverter_price']).number_format = currency_fmt
        ws.cell(row=row, column=3, value=entry['inverter_gst']).number_format = currency_fmt
        ws.cell(row=row, column=4, value=entry['inverter_total']).number_format = currency_fmt
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = border
        row += 1

        # Other Components Section
        ws.cell(row=row, column=1, value="OTHER COMPONENTS & SERVICES").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = border

        row += 1
        for item_name, item_cost in entry['other_items'].items():
            label = item_name.replace("_", " ").title()
            ws.cell(row=row, column=1, value=label).border = border
            ws.cell(row=row, column=2, value=item_cost).number_format = currency_fmt
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = border
            row += 1

        # Other subtotal
        ws.cell(row=row, column=1, value="Other Components Subtotal").border = border
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=entry['other_total_before_gst']).number_format = currency_fmt
        ws.cell(row=row, column=3, value=entry['other_gst']).number_format = currency_fmt
        ws.cell(row=row, column=4, value=entry['other_total']).number_format = currency_fmt
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = border
            ws.cell(row=row, column=c).font = Font(bold=True)
        row += 2

        # Grand Totals
        ws.cell(row=row, column=1, value="TOTAL COST (Before GST)").font = total_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=2, value=entry['total_cost_before_gst']).number_format = currency_fmt
        ws.cell(row=row, column=2).font = total_font
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = border
        row += 1

        ws.cell(row=row, column=1, value="TOTAL GST").font = total_font
        ws.cell(row=row, column=3, value=entry['total_gst']).number_format = currency_fmt
        ws.cell(row=row, column=3).font = total_font
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = border
        row += 1

        ws.cell(row=row, column=1, value="TOTAL COST (Incl. GST)").font = Font(bold=True, size=13, color="1A3A5C")
        ws.cell(row=row, column=4, value=entry['total_cost']).number_format = currency_fmt
        ws.cell(row=row, column=4).font = Font(bold=True, size=13, color="1A3A5C")
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            ws.cell(row=row, column=c).border = border
        row += 2

        # Profit Analysis
        ws.cell(row=row, column=1, value="PROFIT ANALYSIS").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = border

        row += 1
        ws.cell(row=row, column=1, value="Selling Price (Quote to Customer)").border = border
        ws.cell(row=row, column=4, value=entry['selling_price']).number_format = currency_fmt
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = border
        row += 1

        ws.cell(row=row, column=1, value="Total Cost").border = border
        ws.cell(row=row, column=4, value=entry['total_cost']).number_format = currency_fmt
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = border
        row += 1

        profit_fill = profit_fill_green if entry['profit'] >= 0 else profit_fill_red
        ws.cell(row=row, column=1, value="PROFIT / LOSS").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=entry['profit']).number_format = currency_fmt
        ws.cell(row=row, column=4).font = Font(bold=True, size=12)
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = profit_fill
            ws.cell(row=row, column=c).border = border
        row += 1

        ws.cell(row=row, column=1, value="Profit Margin (%)").border = border
        ws.cell(row=row, column=4, value=entry['margin_pct']).number_format = '0.0"%"'
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = border

    # ────────────────────────────────────────
    # Sheet 2: Summary (if multiple entries)
    # ────────────────────────────────────────
    if len(entries) > 1:
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 14
        ws2.column_dimensions['D'].width = 14
        ws2.column_dimensions['E'].width = 14
        ws2.column_dimensions['F'].width = 14
        ws2.column_dimensions['G'].width = 12

        headers = ["System", "Type", "Total Cost", "Selling Price", "Profit", "Margin %"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for i, entry in enumerate(entries):
            row = i + 2
            ws2.cell(row=row, column=1, value=f"{entry['kw']}kW {entry['phase']}").border = border
            ws2.cell(row=row, column=2, value=entry['customer_type'].title()).border = border
            ws2.cell(row=row, column=3, value=entry['total_cost']).number_format = currency_fmt
            ws2.cell(row=row, column=3).border = border
            ws2.cell(row=row, column=4, value=entry['selling_price']).number_format = currency_fmt
            ws2.cell(row=row, column=4).border = border
            ws2.cell(row=row, column=5, value=entry['profit']).number_format = currency_fmt
            ws2.cell(row=row, column=5).border = border
            pf = profit_fill_green if entry['profit'] >= 0 else profit_fill_red
            ws2.cell(row=row, column=5).fill = pf
            ws2.cell(row=row, column=6, value=entry['margin_pct']).number_format = '0.0"%"'
            ws2.cell(row=row, column=6).border = border

        # Totals row
        total_row = len(entries) + 2
        ws2.cell(row=total_row, column=1, value="TOTAL").font = total_font
        ws2.cell(row=total_row, column=3, value=sum(e['total_cost'] for e in entries)).number_format = currency_fmt
        ws2.cell(row=total_row, column=3).font = total_font
        ws2.cell(row=total_row, column=4, value=sum(e['selling_price'] for e in entries)).number_format = currency_fmt
        ws2.cell(row=total_row, column=4).font = total_font
        ws2.cell(row=total_row, column=5, value=sum(e['profit'] for e in entries)).number_format = currency_fmt
        ws2.cell(row=total_row, column=5).font = total_font
        for c in range(1, 7):
            ws2.cell(row=total_row, column=c).fill = section_fill
            ws2.cell(row=total_row, column=c).border = border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
